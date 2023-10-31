from PyQt5.QtWidgets import *
import openpyxl
import sys
import math


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Уравнение')

        self.input_a = QLineEdit()
        self.input_a.setPlaceholderText('a')

        self.input_b = QLineEdit()
        self.input_b.setPlaceholderText('b')

        self.input_x_min = QLineEdit()
        self.input_x_min.setPlaceholderText('x min')

        self.input_x_max = QLineEdit()
        self.input_x_max.setPlaceholderText('x max')

        self.input_dx = QLineEdit()
        self.input_dx.setPlaceholderText('dx')

        self.printer = QPushButton('Вывести')
        self.printer.clicked.connect(self.count_equation)

        self.exporter = QPushButton('Excel')
        self.exporter.clicked.connect(self.export_to_excel)

        self.base_table = QTableWidget()

        # Расположение виджетов
        layout = QVBoxLayout()

        layout.addWidget(self.input_a)
        layout.addWidget(self.input_b)
        layout.addWidget(self.input_x_min)
        layout.addWidget(self.input_x_max)
        layout.addWidget(self.input_dx)

        layout.addWidget(self.printer)
        layout.addWidget(self.exporter)
        layout.addWidget(self.base_table)

        self.setLayout(layout)

    def count_equation(self):
        a = float(self.input_a.text())
        b = float(self.input_b.text())
        x_min = float(self.input_x_min.text())
        x_max = float(self.input_x_max.text())
        dx = float(self.input_dx.text())
        num_steps = int((x_max - x_min) / dx) + 1

        self.base_table.setRowCount(num_steps)
        self.base_table.setColumnCount(2)
        self.base_table.setHorizontalHeaderLabels(["x", "y"])
        self.base_table.verticalHeader().setVisible(False)
        self.base_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.base_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        x = x_min

        for i in range(num_steps):
            y = abs(x**2 - a**2) + ((9*(x**3) / (b - x)**3) * math.sin(x / a))
            y = 'Zero Devision'

            self.base_table.setItem(i, 0, QTableWidgetItem(str(x)))
            self.base_table.setItem(i, 1, QTableWidgetItem(str(round(y, 2)) if type(y) is not str else str(y)))

            x += dx


    def export_to_excel(self):
        try:
            a = float(self.input_a.text())
            b = float(self.input_b.text())
            x_min = float(self.input_x_min.text())
            x_max = float(self.input_x_max.text())
            dx = float(self.input_dx.text())

            num_steps = int((x_max - x_min) / dx) + 1

            wb = openpyxl.Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value='x')
            ws.cell(row=1, column=2, value='y')

            x = x_min
            for i in range(num_steps):
                try:
                    try:
                        y = abs(x ** 2 - a ** 2) + ((9 * (x ** 3) / (b - x) ** 3) * math.sin(x / a))
                    except ZeroDivisionError:
                        y = 'Zero Devision'

                    ws.cell(row=i + 2, column=1, value=x)
                    ws.cell(row=i + 2, column=2, value=(round(y, 2) if type(y) != str else y))
                    x += dx

                except Exception as e:
                    QMessageBox.about(self, "Ошибка", "Введите корректное значение")
                    print(e)

            filename = 'equation.xlsx'
            wb.save(filename)

        except Exception as e:
            QMessageBox.about(self, "Ошибка", "Введите корректное значение")
            print(e)


def main():
    app = QApplication(sys.argv)
    mw = MainWindow()
    mw.show()
    sys.exit(app.exec_())


main()
